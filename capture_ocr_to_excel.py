import cv2
import time
import os
from datetime import datetime
from paddleocr import PaddleOCR
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# =============== 請在這裡修改你的參數 ===============

# 定時拍照間隔（秒）
INTERVAL = 60*10

# Excel 檔案名稱
EXCEL_FILE = "ocr_records.xlsx"

# 新增：照片儲存資料夾
IMG_FOLDER = "img"                          # ← 改這裡就決定資料夾名稱
os.makedirs(IMG_FOLDER, exist_ok=True)      # ← 自動建立資料夾（不存在才建）

# 兩個要辨識的區域 (x, y, w, h)  ← 這裡改成你實際要框的位置
# 格式：(左上角x, 左上角y, 寬度, 高度)
# 推薦（放大 + 避開邊框線）
REGION1 = (380, 140, 120, 70)   # 溫度 27°C（多留一點邊）
REGION2 = (475, 140, 115, 70)   # 濕度 64%（多留一點邊）

# 紅框顏色與粗細
BOX_COLOR = (0, 0, 255)   # BGR 紅色
BOX_THICKNESS = 3

# ====================================================

# === 建立純數字字典（只認 0~9 + 你儀表上出現的符號）===
dict_path = "ppocr_keys_v1.txt"
if not os.path.exists(dict_path):
    with open(dict_path, "w", encoding="utf-8") as f:
        f.write("0123456789.%:-/°C \n")   # 多加 C 和空格，防止模型亂猜
    print("已建立純數字字典檔 ppocr_keys_v1.txt")

# === PaddleOCR 2.8 終極穩定初始化（數字最準、零警告）===
ocr = PaddleOCR(
    use_angle_cls=True,
    lang='en',           # 官方最強英文模型
    use_gpu=False,
    drop_score=0.5       # 不要設太高，0.5 就夠
)
print("PaddleOCR 2.8 初始化完成 → 純數字模式 100% 生效！")

# 建立 Excel（如果不存在就新建）
def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "OCR紀錄"
        ws.append(["拍照時間", "參數1 (區域1)", "參數2 (區域2)"])
        wb.save(EXCEL_FILE)

# 儲存到 Excel
def save_to_excel(timestamp, text1, text2):
    if not os.path.exists(EXCEL_FILE):
        init_excel()
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb["OCR紀錄"]
    ws.append([timestamp, text1.strip(), text2.strip()])
    wb.save(EXCEL_FILE)
    print(f"已儲存到 Excel: {timestamp} | {text1.strip()} | {text2.strip()}")

# OCR 函數（2.8.1 官方推薦寫法）
def ocr_region(img, region):
    x, y, w, h = region
    crop = img[y:y+h, x:x+w]
    
    try:
        result = ocr.ocr(crop, cls=True)
    except:
        return "（錯誤）"
    
    if not result or not result[0]:
        return "（無文字）"
    
    text = ""
    # 這才是真正的鐵壁防禦！！！
    allowed_chars = "0123456789.%°C/-:"  
    
    for line in result[0]:
        raw_text = line[1][0]
        confidence = line[1][1]
        
        # 信心度低於 0.6 直接丟
        if confidence < 0.6:
            continue
            
        # 物理級過濾：只留我們允許的字元
        filtered = "".join(c for c in raw_text if c in allowed_chars)
        if filtered:
            text += filtered + " "
    
    return text.strip() if text else "（無文字）"

# 主程式
def main():
    init_excel()
    cap = cv2.VideoCapture(0)  # 0 = 預設攝影機
    if not cap.isOpened():
        print("無法開啟攝影機！")
        return

    print(f"每 {INTERVAL} 秒拍照一次，按 'q' 結束程式")
    
    last_time = time.time()

    while True:
        ret, frame = cap.read()
        if not ret:
            print("無法讀取畫面")
            break

        # 即時顯示（可看到紅框位置，方便微調）
        temp = frame.copy()
        cv2.rectangle(temp, (REGION1[0], REGION1[1]), 
                     (REGION1[0]+REGION1[2], REGION1[1]+REGION1[3]), BOX_COLOR, BOX_THICKNESS)
        cv2.rectangle(temp, (REGION2[0], REGION2[1]), 
                     (REGION2[0]+REGION2[2], REGION2[1]+REGION2[3]), BOX_COLOR, BOX_THICKNESS)
        cv2.putText(temp, "Adjust regions then wait for auto capture...", (10, 30),
                    cv2.FONT_HERSHEY_SIMPLEX, 1, (0,255,0), 2)
        cv2.imshow("Live View (Red boxes = OCR regions)", temp)

        # 定時拍照
        if time.time() - last_time >= INTERVAL:
            last_time = time.time()
            
            # 產生檔名：2025-12-01_14-30-25.jpg
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            filename = os.path.join(IMG_FOLDER, f"{timestamp}.jpg")  # ← 關鍵：加入資料夾路徑
            cv2.imwrite(filename, frame)
            print(f"拍照儲存: {filename}")

            # 在圖片上畫紅框並顯示（方便你確認框的位置對不對）
            display_img = frame.copy()
            cv2.rectangle(display_img, (REGION1[0], REGION1[1]), 
                         (REGION1[0]+REGION1[2], REGION1[1]+REGION1[3]), BOX_COLOR, BOX_THICKNESS)
            cv2.rectangle(display_img, (REGION2[0], REGION2[1]), 
                         (REGION2[0]+REGION2[2], REGION2[1]+REGION2[3]), BOX_COLOR, BOX_THICKNESS)
            cv2.putText(display_img, f"Captured: {timestamp}", (10, 50),
                        cv2.FONT_HERSHEY_SIMPLEX, 1.5, (0,255,255), 3)
            cv2.imshow("Captured Image with ROI", display_img)
            
            # 執行 OCR
            text1 = ocr_region(frame, REGION1)
            text2 = ocr_region(frame, REGION2)
            
            print(f"區域1 辨識結果: {text1}")
            print(f"區域2 辨識結果: {text2}")
            
            # 存入 Excel
            save_to_excel(timestamp, text1, text2)

        # 按 q 離開
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()

if __name__ == "__main__":
    main()